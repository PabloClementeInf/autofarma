import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
import pandas as pd
from datetime import datetime
import os
import logging
from typing import Dict, List, Any

class ExcelManager:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.output_dir = "./output"
        self.ensure_output_dir()
        
    def ensure_output_dir(self):
        """Crear directorio de salida si no existe"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def is_ready(self) -> bool:
        """Verificar si el manager está listo"""
        return os.path.exists(self.output_dir)
    
    def insert_row_data(self, file_path: str, row_data: List[Any], sheet_name: str = None) -> Dict:
        """Insertar una serie de datos en una nueva fila"""
        try:
            # Verificar si el archivo existe
            if os.path.exists(file_path):
                workbook = openpyxl.load_workbook(file_path)
            else:
                workbook = openpyxl.Workbook()
                
            # Seleccionar hoja
            if sheet_name and sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
                if sheet_name:
                    worksheet.title = sheet_name
            
            # Encontrar la próxima fila vacía
            next_row = worksheet.max_row + 1
            
            # Insertar datos
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=next_row, column=col_num)
                cell.value = value
                
                # Aplicar formato básico
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                elif isinstance(value, datetime):
                    cell.number_format = 'DD/MM/YYYY HH:MM'
            
            # Guardar archivo
            workbook.save(file_path)
            
            return {
                "success": True,
                "message": f"Datos insertados en fila {next_row}",
                "file_path": file_path,
                "row_number": next_row
            }
            
        except Exception as e:
            self.logger.error(f"Error insertando datos en Excel: {e}")
            return {"success": False, "error": str(e)}
    
    def create_pharmacy_report(self, data: List[Dict], report_name: str = None) -> Dict:
        """Crear un reporte específico para farmacia"""
        try:
            if not report_name:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                report_name = f"reporte_farmacia_{timestamp}.xlsx"
            
            file_path = os.path.join(self.output_dir, report_name)
            
            # Crear DataFrame
            df = pd.DataFrame(data)
            
            # Crear archivo Excel con formato
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Reporte', index=False)
                
                # Obtener workbook y worksheet para formatear
                workbook = writer.book
                worksheet = writer.sheets['Reporte']
                
                # Formatear encabezados
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                
                # Ajustar ancho de columnas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            return {
                "success": True,
                "message": "Reporte creado exitosamente",
                "file_path": file_path,
                "records_count": len(data)
            }
            
        except Exception as e:
            self.logger.error(f"Error creando reporte: {e}")
            return {"success": False, "error": str(e)}
    
    def update_inventory_excel(self, product_updates: List[Dict]) -> Dict:
        """Actualizar inventario en Excel"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            inventory_file = os.path.join(self.output_dir, f"inventario_actualizado_{timestamp}.xlsx")
            
            # Crear estructura de datos para inventario
            inventory_data = []
            for update in product_updates:
                inventory_data.append({
                    'Código': update.get('code', ''),
                    'Nombre': update.get('name', ''),
                    'Stock_Anterior': update.get('old_stock', 0),
                    'Stock_Nuevo': update.get('new_stock', 0),
                    'Diferencia': update.get('new_stock', 0) - update.get('old_stock', 0),
                    'Precio': update.get('price', 0),
                    'Proveedor': update.get('supplier', ''),
                    'Fecha_Actualización': datetime.now()
                })
            
            return self.create_pharmacy_report(inventory_data, f"inventario_actualizado_{timestamp}.xlsx")
            
        except Exception as e:
            self.logger.error(f"Error actualizando inventario: {e}")
            return {"success": False, "error": str(e)}